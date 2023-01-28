import re
from functools import cached_property
from openpyxl import Workbook
from typing import Tuple, Protocol, ClassVar

from parsebp.models import (
    PriceList,
    MasterData,
    NewRelease,
    Discontinued,
    References,
)


class UnsupportedFileStructureError(Exception):
    pass


class BPModelTYpe(Protocol):
    sheet_pattern: ClassVar[re.Pattern]


class BoschPrice:
    """
    Represents a loaded Bosch price list.
    """

    def __init__(self, wb: Workbook):
        self.wb = wb
        self.sheet_names = wb.sheetnames

        self.required_model = PriceList
        self.extra_models = [
            MasterData,
            NewRelease,
            Discontinued,
            References,
        ]

    def map_model_to_sheet(self, model: BPModelTYpe) -> Tuple[str, BPModelTYpe] | None:
        """
        Searches in sheet names of the loaded file for patterns of models
        :param model: bp model
        :return: Tuple('Corresponding sheet name', bp model)
        """
        for sheet in self.sheet_names:
            if model.sheet_pattern.search(sheet):
                return sheet, model
        return None

    def mapped_sheets(self):
        if (required_sheet := self.map_model_to_sheet(self.required_model)) is None:
            raise UnsupportedFileStructureError('Unsupported file structure')

        extra_sheets = [
            t for model in self.extra_models
            if (t := self.map_model_to_sheet(model)) is not None
        ]

        return [required_sheet, ] + extra_sheets

    def populate_db(self):
        for sheet, model in self.mapped_sheets():
            ws = self.wb[sheet]
            set_of_values = []
            for i in range(2, ws.max_row + 1):
                data = {
                    field: ws[column['excel_column'] + str(i)]
                    for field, column in model.schema()['properties'].items()
                }
                obj = model(**data)
                set_of_values.append(obj.dict())
